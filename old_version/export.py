import os,sys
import numpy as np
import pickle
from openpyxl import Workbook, load_workbook

# Added because RBS objects now in output of model test system
sys.path.append('../../')
from DNAc import *
from GeneticPartClasses import Class_RNA, Class_Ribosome_Binding_Site

sys.path.append('models')
from RBS_Calculator_v2_0_mod import RBS_Calculator as RBS_Calculator_v2_0
###


datasetInfo_750IC = {
'Dataset': None,
'Abbreviation': 'abbrev',
'Label': 'xls_ID',
'Reporter protein': 'protein',
'Bacterial species': 'organism',
'16 rRNA': 'rRNA',
'Temperature': 'temp',
"5' UTR": 'fivepUTR',
'Coding DNA Sequence (CDS)': 'CDS',
'Mean fluorescence (au)': 'fluo_mean',
'std fluorescence (au)': 'fluo_std'}

datasetInfo_15933FS = {
'Dataset': None,
'Abbreviation': 'abbrev',
'Label': 'xls_ID',
'Reporter protein': 'gene',
'Bacterial species': 'organism',
'16 rRNA': 'rRNA',
'Temperature': 'temp',
"5' UTR": 'fivepUTR',
'Coding DNA Sequence (CDS)': 'CDS',
'RNA.A': 'RNA.A',
'RNA.B': 'RNA.B',
'RNA': 'RNA',
'RNA.var': 'RNA.var',
'RNA.std': 'RNA.std',
'Protein': 'Protein',
'Protein.var': 'Protein.var',
'Protein.std': 'Protein.std',
'transl.rate': 'transl.rate',
'transl.rate.var': 'transl.rate.var',
'transl.rate.std': 'transl.rate.std'
}

thermoModels = {
'mRNA sequence': 'used_mRNA_sequence',
'dG_mRNA': 'dG_mRNA',
'dG_mRNA:rRNA': 'dG_mRNA_rRNA',
'dG_standby': 'dG_standby',
'dG_spacing': 'dG_spacing',
'dG_start': 'dG_start',
'dG_total': 'dG_total',
'dG_pre-SD': 'dG_UTR_fold',
'dG_SD:antiSD': 'dG_SD_16S_hybrid',
'dG_SD': 'dG_SD',
'dG_indirect': 'dG_indirect',
'dG_direct': 'dG_direct',
'dG_post-footprint': 'dG_after_footprint',
'dG_distortion': 'dG_distortion',
'dG_sliding': 'dG_sliding',
'dG_unfolding': 'dG_unfolding',
'Proportionality constant (K)': 'K',
'dG_apparent': 'dG_apparent',
'ddG': 'ddG',
'TIR': 'TIR',
'TIR fold change error': 'fold_change_error',
    
# Additional terms
'dG_DB_aDB': 'dG_DB_aDB',
'DB_struc': 'DB_struc',
'aDB_struc': 'aDB_struc',    

'dot_bracket_structure': 'dot_bracket_structure',
'5pNterminus_ssRNA': '5pNterminus_ssRNA',
'len_ssRNA': 'len_ssRNA',
'AU_count': 'AU_count',
'spacer_sequence': 'spacer_sequence',
'mRNA_rRNA_hybrid_seq': 'mRNA_rRNA_hybrid_seq',
'most_5p_paired_SD_mRNA': 'most_5p_paired_SD_mRNA',
'most_3p_paired_SD_mRNA': 'most_3p_paired_SD_mRNA',
'aligned_most_5p_paired_SD_mRNA': 'aligned_most_5p_paired_SD_mRNA',
'aligned_most_3p_paired_SD_mRNA': 'aligned_most_3p_paired_SD_mRNA',
'dG_mRNA_post_footprint': 'dG_mRNA_post_footprint',
'dG_footprint_subtracted': 'dG_footprint_subtracted',

'num_standby_site_hairpins': 'num_standby_site_hairpins',
'standby_site_hairpin_seqs': 'standby_site_hairpin_seqs',
'standby_site_hairpin_num_bp': 'standby_site_hairpin_num_bp',
'standby_site_hairpin_seq_len': 'standby_site_hairpin_seq_len',
'dG_standby_site_hairpin': 'dG_standby_site_hairpin',

'SD_hairpin_seqs': 'SD_hairpin_seqs',
'SD_hairpin_num_bp': 'SD_hairpin_num_bp',
'SD_hairpin_seq_len': 'SD_hairpin_seq_len',
'dG_SD_hairpin': 'dG_SD_hairpin',

'SD_start_hairpin_seqs': 'SD_start_hairpin_seqs',
'SD_start_hairpin_num_bp': 'SD_start_hairpin_num_bp',
'SD_start_hairpin_seq_len': 'SD_start_hairpin_seq_len',
'dG_SD_start_hairpin': 'dG_SD_start_hairpin',

'start_hairpin_seqs': 'start_hairpin_seqs',
'start_hairpin_num_bp': 'start_hairpin_num_bp',
'start_hairpin_seq_len': 'start_hairpin_seq_len',
'dG_start_hairpin': 'dG_start_hairpin',

'footprint_hairpin_seqs': 'footprint_hairpin_seqs',
'footprint_hairpin_num_bp': 'footprint_hairpin_num_bp',
'footprint_hairpin_seq_len': 'footprint_hairpin_seq_len',
'dG_footprint_hairpin': 'dG_footprint_hairpin',

'post_footprint_hairpin_seqs': 'post_footprint_hairpin_seqs',
'post_footprint_hairpin_num_bp': 'post_footprint_hairpin_num_bp',
'post_footprint_hairpin_seq_len': 'post_footprint_hairpin_seq_len',
'dG_post_footprint_hairpin': 'dG_post_footprint_hairpin'

}

EMOPECModel = {
'pre sequence': 'preseq',
'SD sequence': 'SD_seq',
'spacer sequence': 'spacer_seq',
'spacer length': 'spacer_length',
'Proportionality constant (K)': 'K',
'Expression': 'Expression',
'fold change error': 'fold_change_error',
'Expression_percent': 'Expression_percent'}

RBSDesignerModel = {
'SD sequence': 'SD',
'dG_SD:aSD': 'dG_SD:aSD',
'spacer length': 'spacer length',
'translation efficiency': 'translation efficiency',
'exposure probability': 'exposure probability',
'mRNA-ribosome probability': 'mRNA-ribosome probability',
'Proportionality constant (K)': 'K',
'Expression': 'Expression',
'fold change error': 'fold_change_error'}

def exportExcel(models,datasets,fileName):
    
    handle = open('mRNA_database/mRNA_database.p','rb')
    db = pickle.load(handle)
    handle.close()
    
    handle = open('fast_model_calcs.pkl','rb')
    model_calcs = pickle.load(handle)
    handle.close()
    
    wb = Workbook()
    #wb = load_workbook(fileName)
    
    #=== Write Table of Contents ===#
    
    ws1 = wb.active
    ws1.title = "Table of Contents"

    _ = ws1.cell(column=1,row=1,value="Models")
    _ = ws1.cell(column=2,row=1,value="Datasets")
    _ = ws1.cell(column=3,row=1,value="Authors")
    _ = ws1.cell(column=4,row=1,value="Title")
    _ = ws1.cell(column=5,row=1,value="Journal")
    _ = ws1.cell(column=6,row=1,value="Abbreviation")
    
    for i in range(len(models)):
        _ = ws1.cell(column=1,row=2+i,value=models[i])
    
    for i in range(len(datasets)):
        
        paper = datasets[i]
        _ = ws1.cell(column=2,row=2+i,value=paper)
        
        # Write detailed information
        _ = ws1.cell(column=3,row=2+i,value=db[paper]['authors'])
        _ = ws1.cell(column=4,row=2+i,value=db[paper]['title'])
        _ = ws1.cell(column=5,row=2+i,value=db[paper]['journal'])
        _ = ws1.cell(column=6,row=2+i,value=db[paper]['abbrev'])
    
    # Overwrite paper to replace Table of Contents:
    paper = datasets[0]
    
    # Iterate over models - New sheet for each model
    for m in models:
        
        #name = 'dG_SD_aSD + dG_spacing'
        
        ws = wb.create_sheet(title=m)
        
        # Write lables
        info = ['Dataset',
                'Abbreviation',
                'Label',
                'Reporter protein',
                'Bacterial species',
                '16 rRNA',
                'Temperature',
                "5' UTR",
                'Coding DNA Sequence (CDS)',
                'Mean fluorescence (au)',
                'std fluorescence (au)']
        
        if paper in ['Kosuri_PNAS_2013','Goodman_Science_2013']:
            info = ['Dataset',
                    'Abbreviation',
                    'Label',
                    'Reporter protein',
                    'Bacterial species',
                    '16 rRNA',
                    'Temperature',
                    "5' UTR",
                    'Coding DNA Sequence (CDS)',
                    'RNA.A',
                    'RNA.B',
                    'RNA',
                    'RNA.var',
                    'RNA.std',
                    'Protein',
                    'Protein.var',
                    'Protein.std',
                    'transl.rate',
                    'transl.rate.var',
                    'transl.rate.std']
        
        c = 1
        for k in info:
            _ = ws.cell(column=c,row=1,value=k)
            c += 1
        
        calcs = ['mRNA sequence','dG_mRNA','dG_mRNA:rRNA','dG_SD','dG_standby','dG_spacing','dG_start','dG_indirect','dG_direct',
        'dG_total','dG_pre-SD','dG_SD:antiSD','dG_post-footprint','dG_distortion','dG_sliding',
        'dG_unfolding','dG_DB_aDB','DB_struc','aDB_struc','Proportionality constant (K)','dG_apparent','ddG','TIR','TIR fold change error',
        
        'dot_bracket_structure',
        '5pNterminus_ssRNA',
        'len_ssRNA',
        'AU_count',
        'spacer_sequence',
        'mRNA_rRNA_hybrid_seq',
        'most_5p_paired_SD_mRNA',
        'most_3p_paired_SD_mRNA',
        'aligned_most_5p_paired_SD_mRNA',
        'aligned_most_3p_paired_SD_mRNA',
        'dG_mRNA_post_footprint',
        'dG_footprint_subtracted',
        
        'num_standby_site_hairpins',
        'standby_site_hairpin_num_bp',
        'standby_site_hairpin_seq_len',
        'dG_standby_site_hairpin',
        
        'SD_hairpin_seqs',
        'SD_hairpin_num_bp',
        'SD_hairpin_seq_len',
        'dG_SD_hairpin',
        
        'SD_start_hairpin_seqs',
        'SD_start_hairpin_num_bp',
        'SD_start_hairpin_seq_len',
        'dG_SD_start_hairpin',
        
        'start_hairpin_seqs',
        'start_hairpin_num_bp',
        'start_hairpin_seq_len',
        'dG_start_hairpin',
        
        'footprint_hairpin_seqs',
        'footprint_hairpin_num_bp',
        'footprint_hairpin_seq_len',
        'dG_footprint_hairpin',

        'post_footprint_hairpin_seqs',
        'post_footprint_hairpin_num_bp',
        'post_footprint_hairpin_seq_len',
        'dG_post_footprint_hairpin'
        
        ]
        
        EMOPEC = ['pre sequence','SD sequence','spacer sequence','spacer length','Proportionality constant (K)','Expression','fold change error','Expression_percent']
        RBSDesigner = ['SD sequence','dG_SD:aSD','spacer length','translation efficiency','exposure probability','mRNA-ribosome probability','Proportionality constant (K)','Expression','fold change error']
        
        c += 1
        if m in ['RBS Calculator v1.0','RBS Calculator v1.1','RBS Calculator v2.0','RBS Calculator v2.1','UTR Designer','RBS_Calc_v2_UTR_Designer']:
            for k in calcs:
                v = thermoModels[k]
                if v in model_calcs[paper][m].keys():
                    _ = ws.cell(column=c,row=1,value=k)
                    c += 1
        elif m == 'EMOPEC':
            for k in EMOPEC:
                v = EMOPECModel[k]
                if v in model_calcs[paper][m].keys():
                    _ = ws.cell(column=c,row=1,value=k)
                    c += 1                
        elif m == 'RBS Designer':
            for k in RBSDesigner:
                v = RBSDesignerModel[k]
                if v in model_calcs[paper][m].keys():
                    _ = ws.cell(column=c,row=1,value=k)
                    c += 1
        else:
            pass
        
        # Write data & model predictions
        r = 2
        for paper in datasets:
            
            if paper in ['Kosuri_PNAS_2013','Goodman_Science_2013']:
                datasetInfo = datasetInfo_15933FS
            else:
                datasetInfo = datasetInfo_750IC
            
            for n in range(db[paper]['num_subsets']):
                
                for i in range(db[paper]['subset_size'][n]):
                    
                    c = 1
                    for k in info:
                        v = datasetInfo[k]
                        if k == 'Dataset': _ = ws.cell(column=c,row=r,value=paper)
                        elif k == 'Abbreviation': _ = ws.cell(column=c,row=r,value=db[paper][v])
                        else: _ = ws.cell(column=c,row=r,value=db[paper][v][n][i])
                        c += 1
                    
                    c += 1
                    if m in ['RBS Calculator v1.0','RBS Calculator v1.1','RBS Calculator v2.0','RBS Calculator v2.1','UTR Designer','RBS_Calc_v2_UTR_Designer']:
                        #print model_calcs[paper][m]
                        for k in calcs:
                            v = thermoModels[k]
                            if v in model_calcs[paper][m].keys():
                                if v == 'K': string2write = model_calcs[paper][m][v][n]
                                else: string2write = model_calcs[paper][m][v][n][i]
                                #print string2write
                                _ = ws.cell(column=c,row=r,value=string2write)
                                c += 1
                    elif m == 'EMOPEC':
                        for k in EMOPEC:
                            v = EMOPECModel[k]
                            if v in model_calcs[paper][m].keys():
                                if v == 'K': _ = ws.cell(column=c,row=r,value=model_calcs[paper][m][v][n])
                                else:
                                    #print m,paper,v,len(model_calcs[paper][m][v][n]),i
                                    _ = ws.cell(column=c,row=r,value=model_calcs[paper][m][v][n][i])
                                c += 1
                    elif m == 'RBS Designer':
                        
                        #print paper,m,len(model_calcs[paper][m]['calculated'][0]),i,model_calcs[paper][m]['calculated'][0]
                        
                        if model_calcs[paper][m]['calculated'][n][i]==0:
                            pass
                        else:
                            for k in RBSDesigner:
                                v = RBSDesignerModel[k]
                                if v in model_calcs[paper][m].keys():
                                    if v == 'K': _ = ws.cell(column=c,row=r,value=model_calcs[paper][m][v][n])
                                    else:
                                        #print v
                                        _ = ws.cell(column=c,row=r,value=model_calcs[paper][m][v][n][i])
                                    c += 1
                    else:
                        pass
                    
                    r += 1
    
    wb.save(filename=fileName)
    
    return 1
    

if __name__ == "__main__":
    
    models = ['RBS Calculator v1.0',
              'RBS Calculator v1.1',
              'RBS Calculator v2.0',
              'RBS Calculator v2.1',
              'UTR Designer',
              'RBS Designer',
              'EMOPEC']
    
    models = ['RBS Calculator v2.1']
    
    datasets = ['EspahBorujeni_NAR_2013',
                'EspahBorujeni_NAR_2015',
                'EspahBorujeni_JACS_2016',
                'EspahBorujeni_Footprint',
                'Salis_Nat_Biotech_2009',
                'Farasat_MSB_2014',
                'Tian_NAR_2015',
                'Mimee_Cell_Sys_2015',
                'Bonde_NatMethods_IC_2016']
    
    # datasets = ['Bonde_NatMethods_IC_2016']
    # datasets = ['Kosuri_PNAS_2013']
    # datasets = ['Goodman_Science_2013']
    
    exportExcel(models,datasets,fileName="RBSCalcv1_PyVRNA_fixed_startpos.xlsx")
